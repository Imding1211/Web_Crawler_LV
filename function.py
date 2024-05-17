
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import json
import os

#=============================================================================#

def get_info(url):
    
    home_page = '/'.join(url.split('/')[0:3])
    
    driver = webdriver.Chrome()
    driver.get(url)
    
    html = driver.page_source
    
    driver.quit()
    
    soup = BeautifulSoup(html, "lxml")
    
    products_id_url_element = soup.select('a.lv-smart-link.lv-product-card__url')
    products_price_element = soup.select('span.notranslate')
    
    products_id    = []
    products_name  = []
    products_price = []
    products_url   = []
    
    for product_id_url_element, product_price_element in zip(products_id_url_element, products_price_element):
        
        product_id = product_id_url_element["href"].split('/')[-1]
        products_id.append(product_id)
        
        product_name = product_id_url_element.text.replace('\n','').strip()
        products_name.append(product_name)
        
        product_price = product_price_element.text
        products_price.append(product_price)
        
        product_url = home_page + product_id_url_element["href"]
        products_url.append(product_url)
        
        screen_shot_img(product_url)
        
        print(product_id, product_name, product_price, product_url)
        
    return products_id, products_name, products_price, products_url

#=============================================================================#

def screen_shot_img(url):
    
    product_id = url.split('/')[-1]
    path = './image/side_trunk/' + product_id
    
    if not os.path.exists(path):

        os.mkdir(path)
    
        driver = webdriver.Chrome()
        driver.get(url)
    
        html = driver.page_source
    
        driver.quit()
    
        soup = BeautifulSoup(html, "lxml")
    
        products_img_element = soup.select('div.lv-product-page-header__media img')
        
        img_num = 0
        
        for product_img_element in products_img_element:
            
            img_num += 1
            
            try:
                img_url = product_img_element['srcset'].split(',')[0]
                
            except:
                try:
                    img_url = product_img_element['data-srcset'].split(',')[0]
                    
                except:
                    break
            
            screen_shot_driver = webdriver.Chrome()
            screen_shot_driver.implicitly_wait(3)
            screen_shot_driver.get(img_url)
            screen_shot_driver.save_screenshot(path + '/' + product_id +'-' + str(img_num) + '.png')
            screen_shot_driver.quit()
            
            print(f'Screen shot image {product_id}-{img_num}.')

    else:
        print(f'Image {product_id} already exist.')

#=============================================================================#

def convert_dict(product_data, country, products_id, products_name, products_price, products_url):

    for product_id, product_name, product_price, product_url in zip(products_id, products_name, products_price, products_url):
        
        if product_id not in product_data:

            product_data[product_id] = {}

            product_data[product_id]['img_num']      = count_img_num(product_id)
            product_data[product_id]['product_info'] = {}
            
        if country not in product_data[product_id]:
            product_data[product_id]['product_info'][country] = {}
        
        product_data[product_id]['product_info'][country]['Name']  = product_name
        product_data[product_id]['product_info'][country]['Price'] = product_price
        product_data[product_id]['product_info'][country]['url']   = product_url
            
    return product_data

#=============================================================================#

def count_img_num(product_id):

    path = './image/side_trunk/' + product_id

    file_num = 0

    for lists in os.listdir(path):
        sub_path = os.path.join(path, lists)
        if os.path.isfile(sub_path):
            file_num += 1

    return file_num

#=============================================================================#

def generate_html(product_data):

    headers = ['商品編號', '商品圖片','販售國家', '商品名稱', '商品價格', '商品連結']

    with open('main.html', 'w') as html:
        html.write('<!DOCTYPE html>\n')
        html.write('<html lang="zh-hant-TW">\n')
        html.write('<head>\n')
        html.write('    <meta charset="UTF-8">\n')
        html.write('    <meta name="viewport" content="width=device-width, initial-scale=1.0">\n')
        html.write('    <title>LV Side Trunk</title>\n')
        html.write('    <style>\n')
        html.write('        table {\n')
        html.write('            width: 100%;\n')
        html.write('            border-collapse: collapse;\n')
        html.write('        }\n')
        html.write('        th, td {\n')
        html.write('            border: 1px solid black;\n')
        html.write('            padding: 8px;\n')
        html.write('            text-align: left;\n')
        html.write('        }\n')
        html.write('    </style>\n')
        html.write('</head>\n')
        html.write('<body>\n')
        html.write('    <h1>LV Side Trunk</h1>\n')
        html.write('    <table>\n')
        html.write('        <thead>\n')
        html.write('            <tr>\n')

        for header in headers:
            html.write(f'                <th scope="col">{header}</th>\n')

        for product_id in product_data:
            for index, product_country in enumerate(product_data[product_id]['product_info']):
                html.write('            <tr>\n')

                if index == 0 :
                    html.write(f'                <th rowspan="{len(product_data[product_id]['product_info'])}">{product_id}</th>\n')
                    html.write(f'                <td rowspan="{len(product_data[product_id]['product_info'])}">\n')

                    for num in range(1, product_data[product_id]['img_num']+1):
                        html.write(f'                    <img src="image/side_trunk/{product_id}/{product_id}-{num}.png" style="width:100px;">\n')

                    html.write(f'                </td>\n')

                html.write(f'                <td>{product_country}</td>\n')
                html.write(f'                <td>{product_data[product_id]['product_info'][product_country]['Name']}</td>\n')
                html.write(f'                <td>{product_data[product_id]['product_info'][product_country]['Price']}</td>\n')
                html.write(f'                <td>\n')
                html.write(f'                    <a href="{product_data[product_id]['product_info'][product_country]['url']}">Link</a>\n')
                html.write(f'                </td>\n')
                html.write('            </tr>\n')

        html.write('            </tr>\n')
        html.write('        </thead>\n')
        html.write('        <tbody>\n')
        html.write('            <tr>\n')
        html.write('            </tr>\n')
        html.write('        </tbody>\n')
        html.write('    </table>\n')
        html.write('</body>\n')
        html.write('</html>\n')
        html.write('')

#=============================================================================#

def ouput_excel(product_data, countrys, excel_name):

    heads = ['商品編號', '名稱', '有販售國家', '台灣', '日本', '韓國', '香港', '法國']
    
    workbook = openpyxl.Workbook()
    
    sheet = workbook.worksheets[0]
    
    for index, head in enumerate(heads):
        sheet.cell(1, index+1).value = head
            
    for index, product_id in enumerate(product_data.keys()):
        
        sheet.cell(index+2, 1).value = product_id
        
        sheet.cell(index+2, 3).value = '、'.join(list(product_data[product_id]['product_info']))
        
        for i, country in enumerate(countrys):
            
            if product_data[product_id]['product_info'].get(country) is not None:
                
                product_first_country = list(product_data[product_id]['product_info'])[0]
                
                sheet.cell(index+2, 2).value = product_data[product_id]['product_info'][product_first_country]['Name']
                
                sheet.cell(index+2, i+4).value = product_data[product_id]['product_info'][country]['Price']
                
                sheet.cell(index+2, i+4).hyperlink = product_data[product_id]['product_info'][country]['url']
                sheet.cell(index+2, i+4).style = "Hyperlink"
    
    workbook.save(excel_name)

#=============================================================================#

def ouput_json(product_data, json_name):

    with open(json_name, 'w', encoding='utf-8') as file:
        json.dump(product_data, file, ensure_ascii=False, indent=4)

#=============================================================================#
